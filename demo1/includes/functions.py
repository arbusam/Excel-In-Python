from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from .data import *

def populate_headings(ws, values):
    """Populates the heading for a given worksheet

    Args:
        ws (Worksheet): Worksheet object
        values (List): List of dictionaries containing the headings

    Returns:
        Worksheet: Worksheet object with headings populated
    """

    for i in range(len(values)):
        ws[get_column_letter(i+1) + "1"].value = values[i]["name"]
        ws[get_column_letter(i+1) + "1"].alignment = Alignment(
            horizontal=values[i]["h_align"],
            vertical=values[i]["v_align"]
        )
        ws[get_column_letter(i+1) + "1"].fill = PatternFill(
            fgColor=values[i]["bg_color"], fill_type="solid"
        )
        ws[get_column_letter(i+1) + "1"].font = Font(
            name=values[i]["font_name"],
            size=values[i]["font_size"],
            bold=values[i]["bold"],
            color=values[i]["text_color"],
        )

        ws[get_column_letter(i+1) + "1"].comment = Comment("This is a heading", "System")
        
        ws.column_dimensions[get_column_letter(i+1)].width = values[i]["column_size"]
    return ws

def populate_data(ws, data):
    """Populates the data for a given worksheet

    Args:
        ws (Worksheet): Worksheet object
        data (List): List of lists which each contain the data for one row

    Returns:
        Worksheet: Worksheet object with headings populated
    """

    for row in range(len(data)):
        for column in range(len(data[row])):
            ws[get_column_letter(column+1) + str(row+2)].value = data[row][column]